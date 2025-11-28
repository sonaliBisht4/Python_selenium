import softest
import logging
import inspect
import csv
from openpyxl import load_workbook


class Util(softest.TestCase):

    # ----------------------------------------------------
    # LOGGER
    # ----------------------------------------------------
    @staticmethod
    def custom_logger(level=logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(level)

        formatter = logging.Formatter(
            "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
        )

        if not logger.handlers:
            # Console
            ch = logging.StreamHandler()
            ch.setFormatter(formatter)
            ch.setLevel(level)
            logger.addHandler(ch)

            # File
            fh = logging.FileHandler("test_log.log", mode="a")
            fh.setFormatter(formatter)
            fh.setLevel(level)
            logger.addHandler(fh)

        return logger

    # ----------------------------------------------------
    # PRINT + LOG
    # ----------------------------------------------------
    @staticmethod
    def log_print(logger, msg):
        logger.info(msg)
        print(msg)

    # ----------------------------------------------------
    # SOFT ASSERT FOR FLIGHT LIST
    # ----------------------------------------------------
    def assertListItems(self, elements, expected):
        for el in elements:
            actual = el.get_attribute("aria-label")
            print("Checking:", actual)

            self.soft_assert(
                self.assertEqual,
                actual,
                expected,
                f"FAILED: expected '{expected}' but got '{actual}'"
            )

        self.assert_all()
        print("✔ All assertions passed!")

    # ----------------------------------------------------
    # READ EXCEL
    # ----------------------------------------------------
    @staticmethod
    def readdata_from_excel(path, sheet):
        data = []
        wb = load_workbook(filename=path)
        sh = wb[sheet]

        for r in range(2, sh.max_row + 1):  # skip header
            row = []
            for c in range(1, sh.max_column + 1):
                row.append(sh.cell(row=r, column=c).value)
            data.append(row)

        return data

    # ----------------------------------------------------
    # READ CSV
    # ----------------------------------------------------
    @staticmethod
    def readdata_from_csv(path):
        data = []
        with open(path, "r") as file:
            reader = csv.reader(file)
            next(reader)     # skip header
            for row in reader:
                data.append(row)
        return data
