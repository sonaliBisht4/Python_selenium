import time
import csv
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook

options = Options()
options.add_argument("--start-maximized")

driver = webdriver.Edge(options=options)


workbook = load_workbook("../testdata.xlsx")
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    username, password = row
    driver.get("https://www.saucedemo.com/")
    time.sleep(2)
    driver.find_element(By.ID, "user-name").send_keys(username)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.ID, "login-button").click()

    try:
        # Wait for inventory page
        WebDriverWait(driver, 5).until(EC.url_contains("inventory"))
        print(f" Login successful for: {username}")

        # Click menu and then logout safely
        driver.find_element(By.ID, "react-burger-menu-btn").click()
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "logout_sidebar_link"))
        ).click()
    except NoSuchElementException:
        print(f" Login failed for: {username}")
    time.sleep(2)

driver.quit()
