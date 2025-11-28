import softest
import logging
import inspect
import csv
from openpyxl import load_workbook


class Util(softest.TestCase):

    # -------------------------------
    # LOGGER
    # -------------------------------
    @staticmethod
    def custom_logger(logLevel=logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)

        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )

        # Console handler
        ch = logging.StreamHandler()
        ch.setFormatter(formatter)
        ch.setLevel(logLevel)

        # File handler
        fh = logging.FileHandler("test_log.log", mode="a")
        fh.setFormatter(formatter)
        fh.setLevel(logLevel)

        # Avoid duplicate handlers
        if not logger.handlers:
            logger.addHandler(ch)
            logger.addHandler(fh)

        return logger

    # Print + logger
    @staticmethod
    def log_print(logger, message):
        logger.info(message)
        print(message)

    # -------------------------------
    # SOFT ASSERT LIST
    # -------------------------------
    def assertListItems(self, listitems, expected_value):

        for stop in listitems:
            actual_value = stop.get_attribute("aria-label")
            print("Checking:", actual_value)

            # Soft assertion
            self.soft_assert(
                self.assertEqual,
                actual_value,
                expected_value,
                f"Assertion failed! Expected '{expected_value}' but got '{actual_value}'"
            )

        self.assert_all()
        print("✔ All assertions passed!")

    # -------------------------------
    # EXCEL DATA READER
    # -------------------------------
    @staticmethod
    def readdata_from_excel(file_name, sheet_name):
        datalist = []

        wb = load_workbook(filename=file_name)
        sh = wb[sheet_name]

        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):  # Skip header
            row_data = []
            for j in range(1, col_ct + 1):
                row_data.append(sh.cell(row=i, column=j).value)
            datalist.append(row_data)

        return datalist

    # -------------------------------
    # CSV DATA READER
    # -------------------------------
    @staticmethod
    def readdata_from_csv(filename):
        rows = []
        with open(filename, "r") as file:
            csvreader = csv.reader(file)
            next(csvreader)  # skip header row
            for row in csvreader:
                rows.append(row)
        return rows
